"""Excel export for economics scenario comparisons."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from haulpave.economics.compare import ComparisonResult

__all__ = ["export_comparison_to_excel"]


def export_comparison_to_excel(comparison: ComparisonResult, filepath: str | Path) -> str:
    """Export scenario comparison data to an Excel workbook.

    Parameters
    ----------
    comparison:
        The :class:`ComparisonResult` to export.
    filepath:
        Destination path for the ``.xlsx`` file.

    Returns
    -------
    str
        The resolved absolute path of the created file.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scenario Comparison"

    # -- styling helpers --------------------------------------------------
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    money_fmt = "#,##0.00"

    # -- metadata block ----------------------------------------------------
    ws["A1"] = "Method"
    ws["A1"].font = header_font
    ws["B1"] = comparison.method

    ws["A2"] = "Confidence"
    ws["A2"].font = header_font
    ws["B2"] = comparison.confidence

    ws["A3"] = "Generated At"
    ws["A3"].font = header_font
    from datetime import datetime, timezone

    ws["B3"] = datetime.now(timezone.utc).isoformat(timespec="seconds")

    # -- header row --------------------------------------------------------
    headers = [
        "Scenario",
        "Fuel Cost [USD/year]",
        "Tyre Cost [USD/year]",
        "Maintenance Cost [USD/year]",
        "Total Cost [USD/year]",
    ]
    header_row = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # -- data rows ---------------------------------------------------------
    for i, sc in enumerate(comparison.scenarios, start=1):
        row = header_row + i
        total = (
            sc.fuel_cost_usd_per_year + sc.tire_cost_usd_per_year + sc.maintenance_cost_usd_per_year
        )

        ws.cell(row=row, column=1, value=sc.name)
        c2 = ws.cell(row=row, column=2, value=round(sc.fuel_cost_usd_per_year, 2))
        c2.number_format = money_fmt
        c3 = ws.cell(row=row, column=3, value=round(sc.tire_cost_usd_per_year, 2))
        c3.number_format = money_fmt
        c4 = ws.cell(row=row, column=4, value=round(sc.maintenance_cost_usd_per_year, 2))
        c4.number_format = money_fmt
        cell_total = ws.cell(row=row, column=5, value=round(total, 2))
        cell_total.number_format = money_fmt
        cell_total.font = total_font

    # -- auto-width columns (approximate) -----------------------------------
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=header_row, column=col_idx).value))
        for row_idx in range(header_row + 1, header_row + 1 + len(comparison.scenarios)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(f"${val:,.2f}" if isinstance(val, float) else str(val)))
        ws.column_dimensions[chr(64 + col_idx)].width = max_len + 3

    wb.save(filepath)
    return str(Path(filepath).resolve())
